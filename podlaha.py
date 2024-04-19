import openpyxl
import argparse
import os
import errno
from openpyxl.styles import PatternFill

parser = argparse.ArgumentParser()
parser.add_argument("--nazovsuboru", default="podlaha.xlsx", type=str, help="názov súboru.")


def save_workbook(workbook: openpyxl.workbook.workbook.Workbook, cesta_k_suboru: str) -> int:
    """
    This function tries to save the given workbook to the given file path.
    If the file cannot be saved, it returns an error code.

    Parameters
    ----------
    workbook : openpyxl.workbook.workbook.Workbook
        The workbook object to save.
    cesta_k_suboru : str
        The file path to save the workbook to.

    Returns
    -------
    int
        0 if the workbook was saved successfully, 1 otherwise.
    """
    try:
        # Try to save the workbook
        workbook.save(cesta_k_suboru)
    except Exception as e:
        # If an error occurs, print the error message and return 1
        print(f"Chyba pri ukladaní súboru: {e}")
        return 1

    # If the workbook was saved successfully, return 0
    return 0

def get_sheet(cesta_k_suboru: str, nazov_harka: str) -> (openpyxl.workbook.workbook.Workbook, openpyxl.worksheet.worksheet.Worksheet):
    """
    This function checks if the file exists, and if it does, it tries to open it.
    If the file is open, it raises a FileNotFoundError.
    If the sheet with the given name exists, it removes it.
    Then, it creates a new sheet with the given name and returns it.

    Parameters
    ----------
    cesta_k_suboru : str
        The file path to the Excel file.
    nazov_harka : str
        The name of the sheet in the Excel file.

    Returns
    -------
    openpyxl.workbook.workbook.Workbook
        The opened or created workbook object.
    openpyxl.worksheet.worksheet.Worksheet
        The created sheet object.

    Raises
    ------
    FileNotFoundError
        If the file is open in another application.
    """

    # Check if the file exists
    if not os.path.exists(cesta_k_suboru):
        # If the file doesn't exist, create a new workbook
        workbook = openpyxl.Workbook()
    else:
        # If the file exists, try to open it
        try:
            workbook = openpyxl.load_workbook(cesta_k_suboru)
        except PermissionError:
            # If a PermissionError is raised, print an error message and return None
            print("Súbor je momentálne otvorený v inej aplikácii. Skúste ho zatvoriť a skúste to znova.")
            return None, None

    # Check if the sheet with the given name exists
    if nazov_harka in workbook.sheetnames:
        # If the sheet exists, remove it
        workbook.remove(workbook[nazov_harka])

    # Create a new sheet with the given name
    sheet = workbook.create_sheet(nazov_harka)

    # Return the opened or created workbook object and the created sheet object
    return workbook, sheet


def modeluj_obklad_s_otvorom(
    sirka_podlahy: int,
    dlzka_podlahy: int,
    sirka_obkladacky: int,
    dlzka_obkladacky: int,
    obkladacka_fill: PatternFill,
    velkost_spar: int,
    hrubka_spar: float,
    hustota: float,
    spar_fill: PatternFill,
    cesta_k_suboru: str,
    nazov_harka: str,
    otvor_x: int,
    otvor_y: int,
    otvor_sirka: int,
    otvor_dlzka: int,
    otvor_fill: PatternFill
) -> int:
    """
    This function models a floor or wall with tiles and grout, and calculates the total area, volume, and weight of the grout.
    It also handles the case where there is a hole/opening in the floor/wall.

    Parameters
    ----------
    sirka_podlahy : int
        The width of the floor/wall.
    dlzka_podlahy : int
        The length of the floor/wall.
    sirka_obkladacky : int
        The width of a single tile.
    dlzka_obkladacky : int
        The length of a single tile.
    obkladacka_fill : PatternFill
        The fill color for the tiles.
    velkost_spar : int
        The size of the grout between the tiles.
    hrubka_spar : float
        The thickness of the grout.
    hustota : float
        The density of the grout material.
    spar_fill : PatternFill
        The fill color for the grout.
    cesta_k_suboru : str
        The file path to the Excel file.
    nazov_harka : str
        The name of the sheet in the Excel file.
    otvor_x : int
        The x-coordinate of the hole/opening in the floor/wall.
    otvor_y : int
        The y-coordinate of the hole/opening in the floor/wall.
    otvor_sirka : int
        The width of the hole/opening.
    otvor_dlzka : int
        The length of the hole/opening.
    otvor_fill : PatternFill
        The fill color for the hole/opening.

    Returns
    -------
    0 : The function executed successfully.
    1 : The sheet was not created successfully.
    2 : The width or length of the floor/wall is not positive.
    3 : The width or length of the tile is not positive.
    4 : The size, thickness, or density of the grout is not positive.
    5 : The fill color for the tiles or grout is not a valid PatternFill object.
    6 : The width or length of the hole/opening is not positive.
    7 : The hole/opening is not within the bounds of the floor/wall.
    """
    # Check that the input parameters are valid
    if sirka_podlahy <= 0 or dlzka_podlahy <= 0:
        # If the width or length of the floor/wall is not positive, return an error code
        return 2

    if sirka_obkladacky <= 0 or dlzka_obkladacky <= 0:
        # If the width or length of the tile is not positive, return an error code
        return 3

    if velkost_spar <= 0 or hrubka_spar <= 0 or hustota <= 0:
        # If the size, thickness, or density of the grout is not positive, return an error code
        return 4
    
    if not isinstance(obkladacka_fill, PatternFill) or not isinstance(spar_fill, PatternFill) or not isinstance(otvor_fill, PatternFill):
        # If the fill color for the tiles or grout is not a valid PatternFill object, return an error code
        return 5

    if otvor_sirka <= 0 or otvor_dlzka <= 0:
        # If the width or length of the hole/opening is not positive, return an error code
        return 6

    if otvor_x < 0 or otvor_y < 0 or otvor_x + otvor_sirka > sirka_podlahy or otvor_y + otvor_dlzka > dlzka_podlahy:
        # If the hole/opening is not within the bounds of the floor/wall, return an error code
        return 7
    
    # otvoríme excelovský súbor a hárok
    workbook, sheet = get_sheet(cesta_k_suboru, nazov_harka)

    # Check if the sheet was created successfully
    if workbook is None or sheet is None:
        # If the sheet was not created successfully, return 1
        return 1

    # Inicializujeme pozície a veľkosti
    sirka_stlpca = sirka_obkladacky/10
    vyska_riadka = dlzka_obkladacky/2
    i = 1
    sirka = 0

    # Nastavíme veľkosť stĺpcov podľa obkladu a špáry - horný riadok
    while sirka < sirka_podlahy:
        if i % 2 == 0:  # obkladačka
            if sirka + sirka_obkladacky + velkost_spar <= sirka_podlahy:
                # Ak sa obkladačka a špára zmestia, nastavíme plnú veľkosť obkladačky
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = sirka_stlpca
                sirka += sirka_obkladacky
                sheet.cell(row=1, column=i).value = sirka_obkladacky 
            else:
                # Ak sa obkladačka a špára nezmestia, nastavíme zvyšnú veľkosť obkladačky
                orezana_sirka = sirka_podlahy - sirka - velkost_spar
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = orezana_sirka / (sirka_obkladacky/sirka_stlpca)
                sirka = sirka_podlahy - velkost_spar
                sheet.cell(row=1, column=i).value = orezana_sirka  
        else:  # špára
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = velkost_spar #velkost_spar
            sheet.cell(row=1, column=i).value = velkost_spar 
            sirka += velkost_spar
        i += 1

    # Po urobení prvého radu (stĺpca) robíme prvý rad (riadok)
    j = 1
    dlzka = 0
    while dlzka < dlzka_podlahy:
        if j % 2 == 0:  # obkladačka
            if dlzka + dlzka_obkladacky + velkost_spar <= dlzka_podlahy:
                # Ak sa obkladačka a špára zmestia, nastavíme plnú veľkosť obkladačky
                sheet.row_dimensions[j].height = vyska_riadka
                dlzka += dlzka_obkladacky
                sheet.cell(row=j, column=1).value = dlzka_obkladacky
            else:
                # Ak sa obkladačka a špára nezmestia, nastavíme zvyšnú veľkosť obkladačky
                orezana_dlzka = dlzka_podlahy - dlzka - velkost_spar
                sheet.row_dimensions[j].height = orezana_dlzka / (dlzka_obkladacky/dlzka_obkladacky)
                dlzka = dlzka_podlahy - velkost_spar
                sheet.cell(row=j, column=1).value = orezana_dlzka 
        else:  # špára
            sheet.row_dimensions[j].height = 10 # velkost_spar
            sheet.cell(row=j, column=1).value = velkost_spar 
            dlzka += velkost_spar
        j += 1


    obkladacka_cislo = 0
    pozicia_x = 0
    pozicia_y = 0
    diery_riadok = set()
    stlpcove_spary = {}
    pocet_spar = 0


    # Cyklus pre stĺpce a riadky
    for riadok in range(1, j):
        # Získame šírku a výšku aktuálnej bunky
        pozicia_x = 0
        if riadok % 2 == 0:  # obkladačka
            pozicia_y +=sirka_obkladacky
        else:  # špára
            pozicia_y += velkost_spar

        for stlpec in range(1, i):
            # Získame šírku a výšku aktuálnej bunky
            if stlpec % 2 == 0:  # obkladačka
                pozicia_x += dlzka_obkladacky
            else:  # špára
                pozicia_x += velkost_spar
            # Ak je bunka v oblasti otvoru, nastavíme ju na bielu
            if otvor_x <= pozicia_x <= otvor_x + otvor_sirka and otvor_y <= pozicia_y <= otvor_y + otvor_dlzka:
                sheet.cell(row=riadok, column=stlpec).fill =    otvor_fill
                diery_riadok.add(riadok)

            elif stlpec % 2 == 0 and riadok % 2 == 0:  # obkladačka
                sheet.cell(row=riadok, column=stlpec).fill = obkladacka_fill  # Nastavíme farbu obkladačky
                obkladacka_cislo += 1
                sheet.cell(row=riadok, column=stlpec).value = obkladacka_cislo
            else:  # špára
                sheet.cell(row=riadok, column=stlpec).fill = spar_fill  # Nastavíme farbu špáry
                pocet_spar += 1 
        stlpcove_spary[riadok]=pocet_spar
        pocet_spar = 0

             # Inicializujeme premenné pre obsah obkladačky a špáry
    obsah_obkladacky = 0
    obsah_spar = 0

    for riadok in range(1, j):
        if not(riadok % 2 == 0):  # špára
            if riadok in diery_riadok:
                # ak je posledny rad orezany
                obsah_spar += velkost_spar*(sirka_podlahy-otvor_sirka)  # Pridáme obsah špáry
                sheet.cell(row=riadok, column=i).value = velkost_spar*(sirka_podlahy-otvor_sirka)
            else:
                obsah_spar += velkost_spar*sirka_podlahy  # Pridáme obsah špáry
                sheet.cell(row=riadok, column=i).value = velkost_spar*sirka_podlahy
            #print(f"{riadok=},{obsah_spar=}")
        else:
            if riadok >= j-2:
                obsah_spar += stlpcove_spary[riadok]*3*orezana_dlzka
                sheet.cell(row=riadok, column=i).value = stlpcove_spary[riadok]*velkost_spar*orezana_dlzka
            else:
                obsah_spar += stlpcove_spary[riadok]*velkost_spar*dlzka_obkladacky
                sheet.cell(row=riadok, column=i).value = stlpcove_spary[riadok]*velkost_spar*dlzka_obkladacky



    # Zobrazíme sumárne informácie na konci
    sheet.cell(row=j+1, column=1).value = nazov_harka
    sheet.cell(row=j+2, column=1).value = f"Otvor pozícia: {otvor_x}x{otvor_y}, Otvor rozmery: {otvor_sirka}x{otvor_dlzka}"
    sheet.cell(row=j+3, column=1).value = f"Stena: {sirka_podlahy}x{dlzka_podlahy}, Obklad: {sirka_obkladacky}x{dlzka_obkladacky}, Špáry: {velkost_spar}mm x {hrubka_spar}mm, Hustota špárovaceho materiálu: {hustota} kg/liter"
    sheet.cell(row=j+4, column=1).value = f"Počet použitých obkladačiek: {obkladacka_cislo} ks"
    sheet.cell(row=j+4, column=6).value = obkladacka_cislo
    sheet.cell(row=j+5, column=1).value = f"Obsah obkladu: {obkladacka_cislo*sirka_obkladacky*dlzka_obkladacky/1000000} m2"
    sheet.cell(row=j+5, column=8).value = obkladacka_cislo*sirka_obkladacky*dlzka_obkladacky/1000000
    sheet.cell(row=j+6, column=1).value = f"Obsah špár: {obsah_spar/1000000} m2"
    sheet.cell(row=j+6, column=10).value = obsah_spar/1000000
    sheet.cell(row=j+7, column=1).value = f"Objem špár: {obsah_spar*hrubka_spar/1000000} litrov"
    sheet.cell(row=j+7, column=12).value = obsah_spar*hrubka_spar/1000000
    sheet.cell(row=j+8, column=1).value = f"Hmotnosť špárovacej hmoty: {obsah_spar*hrubka_spar*hustota/1000000} kg"
    sheet.cell(row=j+8, column=14).value = obsah_spar*hrubka_spar*hustota/1000000


    # Uložíme zmeny do súboru
    if not workbook is None:
        result = save_workbook(workbook, cesta_k_suboru)
        if result != 0:
            # If the workbook could not be saved, return 1
            return 1
    return 0

def modeluj_obklad(
    sirka_podlahy: int,
    dlzka_podlahy: int,
    sirka_obkladacky: int,
    dlzka_obkladacky: int,
    obkladacka_fill: PatternFill,
    velkost_spar: int,
    hrubka_spar: float,
    hustota: float,
    spar_fill: PatternFill,
    cesta_k_suboru: str,
    nazov_harka: str
) -> int:
    """
    This function models a floor or wall with tiles and grout, and calculates the total area, volume, and weight of the grout.

    Parameters
    ----------
    sirka_podlahy : int
        The width of the floor/wall.
    dlzka_podlahy : int
        The length of the floor/wall.
    sirka_obkladacky : int
        The width of a single tile.
    dlzka_obkladacky : int
        The length of a single tile.
    obkladacka_fill : PatternFill
        The fill color for the tiles.
    velkost_spar : int
        The size of the grout between the tiles.
    hrubka_spar : float
        The thickness of the grout.
    hustota : float
        The density of the grout material.
    spar_fill : PatternFill
        The fill color for the grout.
    cesta_k_suboru : str
        The file path to the Excel file.
    nazov_harka : str
        The name of the sheet in the Excel file.

    Returns
    -------
    0 : The function executed successfully.
    2 : The width or length of the floor/wall is not positive.
    3 : The width or length of the tile is not positive.
    4 : The size, thickness, or density of the grout is not positive.
    5 : The fill color for the tiles or grout is not a valid PatternFill object.
    """
    # Check that the input parameters are valid
    if sirka_podlahy <= 0 or dlzka_podlahy <= 0:
        # If the width or length of the floor/wall is not positive, return an error code
        return 2

    if sirka_obkladacky <= 0 or dlzka_obkladacky <= 0:
        # If the width or length of the tile is not positive, return an error code
        return 3

    if velkost_spar <= 0 or hrubka_spar <= 0 or hustota <= 0:
        # If the size, thickness, or density of the grout is not positive, return an error code
        return 4
    
    if not isinstance(obkladacka_fill, PatternFill) or not isinstance(spar_fill, PatternFill):
        # If the fill color for the tiles or grout is not a valid PatternFill object, return an error code
        return 5

    # otvoríme excelovský súbor a hárok
    workbook, sheet = get_sheet(cesta_k_suboru, nazov_harka)

    # Check if the sheet was created successfully
    if workbook is None or sheet is None:
        # If the sheet was not created successfully, return 1
        return 1

    # Inicializujeme pozície a veľkosti
    sirka_stlpca = sirka_obkladacky/10
    vyska_riadka = dlzka_obkladacky/2
    i = 1
    sirka = 0

    # Nastavíme veľkosť stĺpcov podľa obkladu a špáry
    while sirka < sirka_podlahy:
        if i % 2 == 0:  # obkladačka
            if sirka + sirka_obkladacky + velkost_spar <= sirka_podlahy:
                # Ak sa obkladačka a špára zmestia, nastavíme plnú veľkosť obkladačky
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = sirka_stlpca
                sirka += sirka_obkladacky
                sheet.cell(row=1, column=i).value = sirka_obkladacky 
            else:
                # Ak sa obkladačka a špára nezmestia, nastavíme zvyšnú veľkosť obkladačky
                orezana_sirka = sirka_podlahy - sirka - velkost_spar
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = orezana_sirka / (sirka_obkladacky/sirka_stlpca)
                sirka = sirka_podlahy - velkost_spar
                sheet.cell(row=1, column=i).value = orezana_sirka  
        else:  # špára
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = velkost_spar #velkost_spar
            sheet.cell(row=1, column=i).value = velkost_spar 
            sirka += velkost_spar
        i += 1

    # Po urobení prvého radu (stĺpca) robíme prvý rad (riadok)
    j = 1
    dlzka = 0
    while dlzka < dlzka_podlahy:
        if j % 2 == 0:  # obkladačka
            if dlzka + dlzka_obkladacky + velkost_spar <= dlzka_podlahy:
                # Ak sa obkladačka a špára zmestia, nastavíme plnú veľkosť obkladačky
                sheet.row_dimensions[j].height = vyska_riadka
                dlzka += dlzka_obkladacky
                sheet.cell(row=j, column=1).value = dlzka_obkladacky
            else:
                # Ak sa obkladačka a špára nezmestia, nastavíme zvyšnú veľkosť obkladačky
                orezana_dlzka = dlzka_podlahy - dlzka - velkost_spar
                sheet.row_dimensions[j].height = orezana_dlzka / (dlzka_obkladacky/dlzka_obkladacky)
                dlzka = dlzka_podlahy - velkost_spar
                sheet.cell(row=j, column=1).value = orezana_dlzka 
        else:  # špára
            sheet.row_dimensions[j].height = 10 # velkost_spar
            sheet.cell(row=j, column=1).value = velkost_spar 
            dlzka += velkost_spar
        j += 1

    obkladacka_cislo = 0
    stlpcove_spary = {}
    pocet_spar = 0
    # Cyklus pre stĺpce a riadky
    for riadok in range(1, j):

        for stlpec in range(1, i):
            if stlpec % 2 == 0 and riadok % 2 == 0:  # obkladačka
                sheet.cell(row=riadok, column=stlpec).fill = obkladacka_fill  # Nastavíme farbu obkladačky
                obkladacka_cislo += 1
                sheet.cell(row=riadok, column=stlpec).value = obkladacka_cislo
            else:  # špára
                sheet.cell(row=riadok, column=stlpec).fill = spar_fill  # Nastavíme farbu špáry
                pocet_spar += 1 
        stlpcove_spary[riadok]=pocet_spar
        pocet_spar = 0
    
    # Inicializujeme premenné pre obsah obkladačky a špáry
    obsah_obkladacky = 0
    obsah_spar = 0
    for riadok in range(1, j):
        if not(riadok % 2 == 0):  # špára
            obsah_spar += velkost_spar*sirka_podlahy  # Pridáme obsah špáry
            sheet.cell(row=riadok, column=i).value = velkost_spar*sirka_podlahy
        else:
            # osetrenie posledneho radu s orezanou kachlickou / obkladom
            if riadok >= j - 2:
                obsah_spar += stlpcove_spary[riadok]*velkost_spar*orezana_dlzka
                sheet.cell(row=riadok, column=i).value = stlpcove_spary[riadok]*velkost_spar*orezana_dlzka
            else:
                obsah_spar += stlpcove_spary[riadok]*velkost_spar*dlzka_obkladacky
                sheet.cell(row=riadok, column=i).value = stlpcove_spary[riadok]*velkost_spar*dlzka_obkladacky

    # Zobrazíme sumárne informácie na konci
    sheet.cell(row=j+1, column=1).value = nazov_harka
    sheet.cell(row=j+2, column=1).value = f"Bez otvoru"
    sheet.cell(row=j+3, column=1).value = f"Stena: {sirka_podlahy}x{dlzka_podlahy}, Obklad: {sirka_obkladacky}x{dlzka_obkladacky} , Špáry: {velkost_spar}mm x {hrubka_spar}mm, Hustota špárovaceho materiálu: {hustota} kg/liter"
    sheet.cell(row=j+4, column=1).value = f"Počet použitých obkladačiek: {obkladacka_cislo} ks"
    sheet.cell(row=j+4, column=6).value = obkladacka_cislo
    sheet.cell(row=j+5, column=1).value = f"Obsah obkladu:{obkladacka_cislo*sirka_obkladacky*dlzka_obkladacky/1000000} m2"
    sheet.cell(row=j+5, column=8).value = obkladacka_cislo*sirka_obkladacky*dlzka_obkladacky/1000000
    sheet.cell(row=j+6, column=1).value = f"Obsah špár:{obsah_spar/1000000} m2"
    sheet.cell(row=j+6, column=10).value = obsah_spar/1000000
    sheet.cell(row=j+7, column=1).value = f"Objem špár:{obsah_spar*hrubka_spar/1000000} litrov"
    sheet.cell(row=j+7, column=12).value = obsah_spar*hrubka_spar/1000000
    sheet.cell(row=j+8, column=1).value = f"Hmotnosť špárovacej hmoty:{obsah_spar*hrubka_spar*hustota/1000000} kg"
    sheet.cell(row=j+8, column=14).value = obsah_spar*hrubka_spar*hustota/1000000


    # Uložíme zmeny do súboru
    if not workbook is None:
        result = save_workbook(workbook, cesta_k_suboru)
        if result != 0:
            # If the workbook could not be saved, return 1
            return 1
    return 0

def main(args: argparse.Namespace):
    args = parser.parse_args([] if "__file__" not in globals() else None)
    if modeluj_obklad(4000, 4300, 185, 598, PatternFill(start_color="808080", end_color="808080", fill_type="solid"), 3, 8, 1.52, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), args.nazovsuboru, "podlaha") == 0:
        print("Podlaha úspešne namodelovaná v exceli")
    modeluj_obklad_s_otvorom(4300, 2500, 200, 200, PatternFill(start_color="808080", end_color="808080", fill_type="solid"),3, 6.5, 1.52, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") ,args.nazovsuboru, "stena s oknom",800,1018,2100,1200, PatternFill(start_color="03a1fc", end_color="03a1fc", fill_type="solid"))
    modeluj_obklad(4000, 2500, 200, 200, PatternFill(start_color="808080", end_color="808080", fill_type="solid"), 3, 6.5, 1.64, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), args.nazovsuboru, "stena2")
    modeluj_obklad_s_otvorom(4300, 2500, 200, 200,PatternFill(start_color="808080", end_color="808080", fill_type="solid"), 3, 6.5, 1.52, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), args.nazovsuboru, "stena s dverami",800,0,800,2000,PatternFill(start_color="99534e", end_color="99534e", fill_type="solid"))
    modeluj_obklad(4000, 2500, 200, 200, PatternFill(start_color="808080", end_color="808080", fill_type="solid"), 3, 6.5, 1.52,  PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), args.nazovsuboru, "stena4")


if __name__ == "__main__":
    args = parser.parse_args([] if "__file__" not in globals() else None)
    main(args)
